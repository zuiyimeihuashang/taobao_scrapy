import os
from openpyxl import load_workbook

def create_folders_from_excel(file_path, column_name, root_dir):
    # 打开Excel文件
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]

    # 获取列索引
    column_index = None
    for col in range(1, sheet.max_column + 1):  # 从1开始，因为 openpyxl 列是从1开始的
        if sheet.cell(row=1, column=col).value == column_name:
            column_index = col
            break

    if column_index is None:
        print(f"未找到列名为 '{column_name}' 的列")
        return

    # 获取指定列的数据
    column_data = []
    for row in range(2, sheet.max_row + 1):  # 从第二行开始，跳过表头
        cell_value = sheet.cell(row=row, column=column_index).value
        column_data.append(cell_value)

    # 遍历每个文件夹名称并创建文件夹
    for folder_name in column_data:
        folder_path = os.path.join(root_dir, str(folder_name))  # 使用 str() 将值转换为字符串
        os.makedirs(folder_path, exist_ok=True)
        print(f"创建文件夹: {folder_path}")

if __name__ == "__main__":
    file_path = "C:/Users/杨雨糯/Desktop/22大数据1班信息表(5)(2) (1).xlsx"
    column_name = "姓名"
    root_dir = "C:/Users/杨雨糯/Desktop/data"

    create_folders_from_excel(file_path, column_name, root_dir)