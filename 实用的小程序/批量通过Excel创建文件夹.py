import os
from openpyxl import load_workbook

def create_folders_from_excel(file_path, column_name, root_dir,agent,last):
    # 打开Excel文件
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]

    # 获取列索引
    column_index = None
    for col in range(1, sheet.max_column + 1):  # 从1开始，因为 openpyxl 列是从1开始的
        if sheet.cell(row=1, column=col).value == column_name:
            column_index = col
            break

    if column_index is None:
        print(f"未找到列名为 '{column_name}' 的列")
        return

    # 获取指定列的数据
    column_data = []
    j = agent
    for row in range(2, sheet.max_row + 1):  # 从第二行开始，跳过表头
        if j>last:
            break
        j += 1
        cell_value = sheet.cell(row=row, column=column_index).value
        column_data.append(cell_value)

    # 遍历每个文件夹名称并创建文件夹
    j = agent
    for folder_name in column_data:
        folder_path = os.path.join(root_dir, str(j)+str(folder_name))  # 使用 str()将值转换为字符串，并拼接路径。
        os.makedirs(folder_path, exist_ok=True)
        # 通过对注释的微调可控制生成文件夹下的子文件夹的生成。
        # zi_dirs=["子文件1","子文件2","子文件3","子文件4"]
        # for zi_dir in zi_dirs:
        #     zi_dirpath = os.path.join(folder_path,zi_dir)
        #     os.makedirs(zi_dirpath, exist_ok=True)
        print(f"创建文件夹: {folder_path}")
        j += 1

if __name__ == "__main__":
    file_path = "C:/Users/杨雨糯/Desktop/22大数据1班信息表(5)(2) (1).xlsx" #待提取文件位置
    column_name = "姓名" #提取的列的名称
    root_dir = "G:/IDM文件/新建文件夹" #新文件夹位置
    agent = 1 #开始编号
    last  = 3 #结束编号
    create_folders_from_excel(file_path, column_name, root_dir,agent,last)
